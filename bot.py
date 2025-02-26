import os
import random
import re
import pandas as pd
import openpyxl
from typing import Final
from telegram import Update, PollAnswer, Poll, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, PollAnswerHandler, CallbackQueryHandler, ContextTypes
from collections import defaultdict
import asyncio
from openpyxl import load_workbook, Workbook
import time
from telegram.error import Forbidden,BadRequest, TimedOut
import telegram

TOKEN: Final = '7938454369:AAEEcc9ILmAoVS3S23WD6KQ8FLPSRS4Pvs4'
BOT_USERNAME: Final = '@slizzyy_bot'
# Bot configuration

ALLOWED_GROUP_IDS = [-1001817635995, -1002114430690,-1001817635995]
EXCEL_FILE = 'SYNO5.xlsx'
SCORE_FILE="user_scores.xlsx"
groupsendid = -1002114430690


# Global state variables

quiz_state = {}
correct_users = {}  # Tracks correct answers per user
quiz_scores = {}  # Stores scores per chat_id

selected_poll_count = 0 
selected_quizscore_count=0
active_poll=1 # Number of polls user requested
answers_received = defaultdict(int)  # Tracks how many answers have been received for each user
is_quiz_active = False  # New variable to track if a quiz is active
chat_id = None  # Current chat ID for the quiz
selected_time_limit = 10  # Default time limit
unanswered_poll = 0
cancel_active = False
display_chat=0
Quiz_grammar_type =''
quiz_kick= False
StudyStuffgrp=False
commandfunctionpass = 1
def get_chat_state(chat_id):
    if chat_id not in quiz_state:
        quiz_state[chat_id] = {
            "is_active": False,
            "selected_polls": [],
            "correct_users": {},
        }
    return quiz_state[chat_id]

# Load quiz data from Excel
used_srnos = set()
def reset_used_srnos():
    global used_srnos
    used_srnos.clear()
def load_quiz_data(file_path, selected_poll_count):
    global used_srnos
    try:
        df = pd.read_excel(file_path)
        
        # Trim extra spaces in each column where cells are strings
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        # Filter out rows that have already been selected based on `srno`
        unique_rows = df[~df['srno'].isin(used_srnos)]
        
        # If there are fewer unique rows than requested polls, adjust to available rows
        if len(unique_rows) < selected_poll_count:
            print("Not enough unique rows available.")
            selected_poll_count = len(unique_rows)
        
        # Select a random sample of unique rows
        selected_rows = unique_rows.sample(n=selected_poll_count)
        
        # Update used_srnos with newly selected rows
        used_srnos.update(selected_rows['srno'].tolist())
        
        # Process selected rows into polls
        polls = []
        for _, row in selected_rows.iterrows():
            options = [row["option1"], row["option2"], row["option3"], row["option4"]]
            random.shuffle(options) 
            poll = {
                "question": row["question"],
                "options": options,
                "correct_answer": row["answer"],
                "meaning": row.get("meaning", "No meaning provided")  # Ensure meaning is loaded
            }
            polls.append(poll)
            
        return polls
    except Exception as e:
        print(e)
def escape_markdown(text: str) -> str:
    return re.sub(r'([_\*\[\]\(\)~`>#+\-=|{}.!])', r'\\\1', text)

def load_scores():
    if not os.path.exists(SCORE_FILE):
        return []

    workbook = openpyxl.load_workbook(SCORE_FILE)
    sheet = workbook.active

    scores = []
    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header

        user_id = sheet.cell(row=row, column=2).value
        username = sheet.cell(row=row, column=3).value
        score = sheet.cell(row=row, column=4).value
        round = sheet.cell(row=row, column=5).value

        if user_id and username and score and round is not None:
            scores.append((user_id, username, score,round))

    workbook.close()
    return scores

async def start_game_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global is_quiz_active,quiz_state, correct_users,commandfunctionpass, chat_id, unanswered_poll,cancel_active,selected_quizscore_count,quiz_kick
        cancel_active = False
        quiz_kick= False
        commandfunctionpass = 1
        reset_used_srnos()
        chat_id = update.message.chat.id
        
        if chat_id not in quiz_state:
            quiz_state[chat_id] = {
                "is_active": True,
                "active":True,
                "polls": [],
                "scores": {},  # Stores user scores
                "quiz_kick": False,  # Tracks if bot was removed
                "cancel_active": False,  # Track cancellation
                "selected_poll_count": 0,  # Store number of rounds
            }

        print("now quiz ",quiz_state)
        
        
        selected_quizscore_count=0
        
        correct_users.clear()  
        difficulty_keyboard = [
            [InlineKeyboardButton("NDA-CDS", callback_data='type_NDA0')],
            [InlineKeyboardButton("English Grammar", callback_data='type_BASIC')],
            [InlineKeyboardButton("UPSC", callback_data='type_Upsc')],
            [InlineKeyboardButton("Jee and Neet", callback_data='type_Neet')],
        ]
        reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
        try:
            await update.message.chat.send_message('Select the Quiz type:', reply_markup=reply_markup)
        except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
#================================================================Option buttons For Quiz Type 
def Nda_keyboard0():
    return [
        [InlineKeyboardButton("Synonyms", callback_data='difficulty_synonyms_nda')],
        [InlineKeyboardButton("Antonyms", callback_data='difficulty_nda_antonyms')],
        [InlineKeyboardButton("Idiom-Phrase", callback_data='difficulty_idiomphrase_nda')],
        [InlineKeyboardButton("One word Substitute", callback_data='difficulty_nda_ows')],
        [InlineKeyboardButton("🧑‍🦯‍➡️ Next 🧑‍🦯‍➡️", callback_data='type_NDA1')]
    ]        

def Nda_keyboard1():
    return [
        
        [InlineKeyboardButton("Active-passive", callback_data='difficulty_acitvepassive_nda')],
        [InlineKeyboardButton("🏎️  Previous ", callback_data='type_NDA0'),InlineKeyboardButton("Next 🧑‍🦯‍➡️", callback_data='type_NDA2')],
        
    ]
def Nda_keyboard2():
    return [
       [InlineKeyboardButton("Fill in the Blanks", callback_data='difficulty_fillblank_nda')],
        [InlineKeyboardButton("🏎️  Previous", callback_data='type_NDA1'),InlineKeyboardButton("Next 🧑‍🦯‍➡️", callback_data='type_NDA0')]
        
    ]

#================================================================Option buttons For Quiz Type 
async def handle_type_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name

        await query.answer()
        await query.answer()
        difficulty_message = ''

        if query.data == 'type_NDA0':
            reply_markup = InlineKeyboardMarkup(Nda_keyboard0())
            try:
                await query.edit_message_text(f'@{username} selected NDA-CDS Phase 1 \n Select the Grammar Quiz type:', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected NDA-CDS Phase 1 \n Select the Grammar Quiz type:', reply_markup=reply_markup)
        

        elif query.data == 'type_NDA1':
            reply_markup = InlineKeyboardMarkup(Nda_keyboard1())
            try:
                await query.edit_message_text(f'@{username} selected NDA-CDS Phase 2 \n\n Select the Grammar Quiz type:', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected NDA-CDS Phase 2 \n\n Select the Grammar Quiz type:', reply_markup=reply_markup)
        

        elif query.data == 'type_NDA2':
            reply_markup = InlineKeyboardMarkup(Nda_keyboard2())
            try:
                await query.edit_message_text(f'@{username} selected NDA-CDS Phase 2 \n\n Select the Grammar Quiz type:', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected NDA-CDS Phase 2\n Select the Grammar Quiz type:', reply_markup=reply_markup)
        

        elif query.data == 'type_BASIC':
            
            difficulty_keyboard = [
                [InlineKeyboardButton("Synonyms", callback_data='difficulty_synonyms')],
                [InlineKeyboardButton("Antonyms", callback_data='difficulty_antonyms')],
                [InlineKeyboardButton("Spelling Correction", callback_data='difficulty_spellcorr')],
                [InlineKeyboardButton("Daily Life Idioms", callback_data='difficulty_shortIdiom')],
                [InlineKeyboardButton("Sentence Correction", callback_data='difficulty_sentcorr')],
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.edit_message_text(f'@{username} Selected Basic Grammar \n Select the Grammar Quiz type:', reply_markup=reply_markup)   
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected Basic Grammar \n Select the Grammar Quiz type:', reply_markup=reply_markup)
        
        elif query.data == 'type_Upsc':
            
            difficulty_keyboard = [
                [InlineKeyboardButton("History", callback_data='difficulty_upschistory')],
                [InlineKeyboardButton("Science Tech", callback_data='difficulty_upscscience')],
                #[InlineKeyboardButton("GeoGraphy", callback_data='difficulty_UpscGeography')],
                #[InlineKeyboardButton("Sentence Correction", callback_data='difficulty_sentcorr')],
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.edit_message_text(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
        elif query.data == 'type_Neet':
            
            difficulty_keyboard = [
                [InlineKeyboardButton("Chemistry", callback_data='difficulty_neetchemistry')]
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.edit_message_text(f'@{username} Selected Jee and Neet \n Select the Quiz type:', reply_markup=reply_markup)   
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected  Jee and  Neet \n Select the Quiz type:', reply_markup=reply_markup)
        
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
          
# Handle difficulty selection
async def handle_difficulty_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global EXCEL_FILE, Quiz_grammar_type,StudyStuffgrp
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name

        await query.answer()
        difficulty_message = ''

        if query.data == 'difficulty_synonyms':
            StudyStuffgrp = False
            EXCEL_FILE = 'SYNO5.xlsx'
            difficulty_message = "Synonyms"
        elif query.data == 'difficulty_antonyms':
            StudyStuffgrp = False
            EXCEL_FILE = 'Antonyms5.xlsx'
            difficulty_message = "Antonyms"
        elif query.data == 'difficulty_spellcorr':
            StudyStuffgrp = False
            EXCEL_FILE = 'spellCorrection4.xlsx'
            difficulty_message = "Spelling Correction"

        elif query.data == 'difficulty_sentcorr':
            StudyStuffgrp = False
            EXCEL_FILE = 'sentenceCorr4.xlsx'
            difficulty_message = "Sentence Correction"
        elif query.data == 'difficulty_shortIdiom':
            StudyStuffgrp = False
            EXCEL_FILE = 'basic_shortidiom.xlsx'
            difficulty_message = "Daily Life Idioms"
        #==========================For Neet==============================
        elif query.data == 'difficulty_neetchemistry':
            StudyStuffgrp = False
            EXCEL_FILE = 'Neet_Chemistry.xlsx'
            difficulty_message = "Chemistry"
        #==========================For NDA==============================
        elif query.data == 'difficulty_synonyms_nda':
            StudyStuffgrp = False
            EXCEL_FILE = 'Nda_Synonyms_updated.xlsx'
            difficulty_message = "Synonyms"
        elif query.data == 'difficulty_acitvepassive_nda':
            StudyStuffgrp = False
            EXCEL_FILE = 'NDA_active_passive_voice1.xlsx'
            difficulty_message = "Active passive Voice"
        elif query.data == 'difficulty_fillblank_nda':
            StudyStuffgrp = False
            EXCEL_FILE = 'Nda_FillinTheBlanks.xlsx'
            difficulty_message = "Fill in the blanks"
        elif query.data == 'difficulty_idiomphrase_nda':
            StudyStuffgrp = False
            EXCEL_FILE = 'Nda_Idiom.xlsx'
            difficulty_message = "Idiom Phrase"
        elif query.data == 'difficulty_nda_ows':
            StudyStuffgrp = False
            EXCEL_FILE = 'Nda_OneWord.xlsx'
            difficulty_message = "One word Substitution"
        elif query.data == 'difficulty_nda_antonyms':
            StudyStuffgrp = False
            EXCEL_FILE = 'Nda_Antonyms.xlsx'
            difficulty_message = "Antonyms"
#==================================================================
        elif query.data == 'difficulty_upschistory':
            StudyStuffgrp = False
            EXCEL_FILE = 'Upsc_history.xlsx'
            difficulty_message = "History"
        elif query.data == 'difficulty_upscscience':
            StudyStuffgrp = False
            EXCEL_FILE = 'Upsc_ScienceandTech.xlsx'
            difficulty_message = "Science Technology"
        # Edit the message to indicate selection and remove other buttons
        Quiz_grammar_type = difficulty_message
        
        if(Quiz_grammar_type !='Reasoning' and Quiz_grammar_type !='Maths'):
            time_keyboard = [
            
            [InlineKeyboardButton("15 Seconds", callback_data='time_15')],
            [InlineKeyboardButton("20 Seconds", callback_data='time_20')],
            [InlineKeyboardButton("25 Seconds", callback_data='time_25')],
            [InlineKeyboardButton("30 Seconds", callback_data='time_30')],
            
            ]
        else:
            time_keyboard = [
            [InlineKeyboardButton("30 Seconds", callback_data='time_30')],
            [InlineKeyboardButton("45 Seconds", callback_data='time_45')],
            [InlineKeyboardButton("60 Seconds", callback_data='time_60')],
            [InlineKeyboardButton("90 Seconds", callback_data='time_90')],
            ]
        reply_markup = InlineKeyboardMarkup(time_keyboard)
        try:
            await query.edit_message_text(f'@{username} Chooses the {difficulty_message}  \n\n Select the time limit for each poll:', reply_markup=reply_markup)
            
        except (BadRequest, Forbidden, TimedOut) as e:
            await query.message.chat.send_message(f'@{username} Chooses the {difficulty_message}  \n\n Select the time limit for each poll:', reply_markup=reply_markup)
    
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
        
async def handle_time_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global selected_time_limit
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name
        
        await query.answer()
        time_mapping = {
            
            'time_15': 15,
            'time_20': 20,
            'time_25': 25,
            'time_30': 30,
            'time_45': 45,
            'time_60': 60,
            'time_90': 90,
        }
        selected_time_limit = time_mapping.get(query.data, 10)
        
        keyboard = [
            [InlineKeyboardButton("15 Words", callback_data='15')],
            [InlineKeyboardButton("25 Words", callback_data='25')],
            [InlineKeyboardButton("35 Words", callback_data='35')],
            [InlineKeyboardButton("50 Words", callback_data='50')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            await query.edit_message_text(f'@{username} selected {selected_time_limit} second To complete one quiz. \n\n How many rounds?', reply_markup=reply_markup)
            
        except (BadRequest, Forbidden, TimedOut) as e:
            await query.message.chat.send_message(f'@{username} selected {selected_time_limit} second To complete one quiz. \n\n How many rounds?', reply_markup=reply_markup)
    
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)

async def handle_button_click(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global quiz_state
        query = update.callback_query
        chat_id = query.message.chat_id  # Identify which chat is running the quiz
        username = query.from_user.username or query.from_user.first_name
        print("cgat id",chat_id)
        print("quiz state",quiz_state)
        # Ensure quiz state is chat-specific
        if chat_id not in quiz_state:
            print("enter here 1")
            quiz_state[chat_id] = {"active": False}

        if not quiz_state[chat_id]["active"]:
            print("enter here 2")
            await query.answer("Please start a new quiz with /startquiz")
            return

        selected_poll_count =  int(query.data)
        quiz_state[chat_id]["total_rounds"] = selected_poll_count
        quiz_state[chat_id]["active"] = True
        quiz_state[chat_id]["polls"] = []


        await query.edit_message_text(text=f"@{username} selected {selected_poll_count} rounds. Starting the quiz...")

        # Load quiz data for the chat
        selected_polls = load_quiz_data(EXCEL_FILE, selected_poll_count)
        

        # Send polls asynchronously
        asyncio.create_task(send_quiz_polls(chat_id, selected_polls, context))

    except Exception as e:
        print(f"Error in handle_button_click: {e}")

async def send_quiz_polls(chat_id, polls, context):
    for i, poll in enumerate(polls):
        if not quiz_state[chat_id]["active"]:
            print(f"Quiz canceled in {chat_id}, stopping polls.")
            break  # Stop sending polls if quiz is canceled

        try:
            poll_message = await context.bot.send_poll(
                chat_id=chat_id,
                question=f"{i+1}/{len(polls)}: {poll['question']}",
                options=poll['options'],
                is_anonymous=False,
                allows_multiple_answers=False,
                type=Poll.QUIZ,
                correct_option_id=poll['options'].index(poll['correct_answer'])
            )

            # Store poll details
            quiz_state[chat_id]["polls"].append({
                "poll_id": poll_message.poll.id,
                "question": poll["question"],
                "correct_answer": poll["correct_answer"],
                "options": poll["options"],
                "meaning": poll["meaning"],
                "responses": {},
                "poll_message": poll_message
            })

            # Start countdown asynchronously
            await countdown_and_close_poll(chat_id, poll_message, context)

            await asyncio.sleep(2)  # Small delay between polls

        except Exception as e:
            print(f"Error sending poll: {e}")

    # ✅ Call calculate_scores **only if the quiz was completed**
    if quiz_state.get(chat_id, {}).get("active", False):
        await calculate_scores(chat_id, context)

    # ✅ Cleanup: Remove quiz state after scoring (optional)
    quiz_state.pop(chat_id, None)


async def cancel_quiz_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id

    if chat_id in quiz_state and quiz_state[chat_id]["active"]:
        quiz_state[chat_id]["active"] = False
        await context.bot.send_message(chat_id, text="The quiz has been canceled. You can restart with /startquiz.")
    else:
        await context.bot.send_message(chat_id, text="No active quiz to cancel.")

async def countdown_and_close_poll(chat_id, poll_message, context):
    try:
        await asyncio.sleep(15)  # Poll stays open for 15 seconds

        # Stop poll
        await poll_message.stop_poll()

        # Fetch the stored poll data
        for poll in quiz_state[chat_id]["polls"]:
            if poll["poll_id"] == poll_message.poll.id:
                meaning = poll["meaning"]
                if meaning:
                    await context.bot.send_message(chat_id, text=f"Meaning: {meaning}")
                break

    except Exception as e:
        print(f"Error in countdown_and_close_poll: {e}")



final_poll_responses = {}

async def handle_poll_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        answer = update.poll_answer
        user_id = str(answer.user.id)
        username = answer.user.username or answer.user.first_name or user_id
        if answer.user.username:
            username = f"@{answer.user.username}"  # Add @ if username exists
        elif answer.user.first_name:
            username = answer.user.first_name  # Use first name if no username
        else:
            username = str(user_id)
        selected_options = answer.option_ids
        

        # Find which chat this poll belongs to
        for chat_id, chat_quiz in quiz_state.items():
            for poll in chat_quiz["polls"]:
                if poll["poll_id"] == answer.poll_id:
                    correct_answer = poll["correct_answer"]
                    options = poll["options"]

                    selected_answer = options[selected_options[0]]  # Assume single choice
                    if selected_answer == correct_answer:
                        # Initialize chat-specific score tracking
                        if chat_id not in quiz_scores:
                            quiz_scores[chat_id] = {}

                        if user_id not in quiz_scores[chat_id]:
                            quiz_scores[chat_id][user_id] = {"username": username, "score": 0}

                        quiz_scores[chat_id][user_id]["score"] += 1
                     

                    return  # Exit after finding the correct poll

    except Exception as e:
        print(f"Error in handle_poll_answer: {e}")


def update_user_score(chat_id, correct_users):
    """
    Update user scores in an Excel file. If the user exists in the same chat_id, update their score.
    If the user exists in a different chat_id, create a new row.
    """
    try:
        game_round = 1
        # Load existing workbook or create a new one
        try:
            workbook = load_workbook(SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Scores"
            sheet.append(["srno", "chatid", "Idnumber", "Username", "Score", "round"])
        existing_scores = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:  # Ensure row has all required columns
                sr_no, existing_chat_id, user_id, username, score, round = row
                existing_scores[(str(existing_chat_id), str(user_id))] = {
                    "sr_no": sr_no,
                    "chat_id": str(existing_chat_id),
                    "username": username,
                    "score": int(score),
                    "round": int(round)
                }
        for user_id, data in correct_users.items():
            username = data["username"]
            new_score = data["score"]
            key = (str(chat_id), str(user_id))
            if key in existing_scores:
                existing_scores[key]["username"] = username
                existing_scores[key]["score"] += new_score
                existing_scores[key]["round"] += game_round
            else:
                sr_no = len(existing_scores) + 1
                existing_scores[key] = {
                    "sr_no": sr_no,
                    "chat_id": str(chat_id),
                    "username": username,
                    "score": new_score,
                    "round": game_round
                }
        sheet.delete_rows(2, sheet.max_row)
        for (chat_id, user_id), data in existing_scores.items():
            sheet.append([data["sr_no"], chat_id, user_id, data["username"], data["score"], data["round"]])
        workbook.save(SCORE_FILE)
        print("Scores updated successfully.")

    except Exception as e:
        print(f"Error updating scores: {e}")


async def calculate_scores(chat_id, context):
    global groupsendid
    if chat_id not in quiz_scores or not quiz_scores[chat_id]:
        await context.bot.send_message(chat_id, "No one participated in the quiz.")
        return
    update_user_score(chat_id, quiz_scores[chat_id])
    if os.path.exists(SCORE_FILE):
        with open(SCORE_FILE, 'rb') as file:
            await context.bot.send_document(chat_id=groupsendid, document=file)

    # Sort players by score (highest first)
    sorted_scores = sorted(quiz_scores[chat_id].items(), key=lambda x: x[1]["score"], reverse=True)

    # Prepare leaderboard message
    leaderboard = f"🏆 *Quiz Results* 🏆\n\n"

    for rank, (user_id, data) in enumerate(sorted_scores, start=1):
        username = escape_markdown(data["username"])  # Removed version=2
        leaderboard += f"{rank}\\) *{username}* \\- `{data['score']} points`\n"

    # Send leaderboard result
    await context.bot.send_message(chat_id, leaderboard, parse_mode="MarkdownV2")

    # Cleanup scores after displaying results
    quiz_scores.pop(chat_id, None)

async def my_rank(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Returns the total score and rounds played by a user across all groups."""
    try:
        user_id = str(update.message.from_user.id)
        username = update.message.from_user.username or update.message.from_user.first_name or user_id

        # Load Excel file
        try:
            workbook = load_workbook(SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            await update.message.reply_text("No scores found.")
            return

        total_score = 0
        total_rounds = 0
        user_found = False

        # Check all rows for user's total score across all groups
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                _, chat_id, stored_user_id, stored_username, score, rounds = row
                if str(stored_user_id) == user_id:
                    total_score += int(score)
                    total_rounds += int(rounds)
                    user_found = True

        if user_found:
            response = f"🏅 *Your Rank*\n\n👤 *Username:* {username}\n🎯 *Total Score:* {total_score}\n🔄 *Rounds Played:* {total_rounds}\n🧑‍🤝‍🧑 *Total Groups:* {len(set(chat_id for row in sheet.iter_rows(min_row=2, values_only=True) if row[2] == user_id))}"
        else:
            response = "You haven't participated in any quizzes yet."

        await update.message.reply_text(response, parse_mode="Markdown")

    except Exception as e:
        print(f"Error in my_rank: {e}")

async def top1grp_scorer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Returns the top 10 users from the current group along with the total number of participants."""
    try:
        chat_id = str(update.message.chat_id)

        # Load Excel file
        try:
            workbook = load_workbook(SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            await update.message.reply_text("No scores found for this group.")
            return

        group_scores = {}
        total_users = set()

        # Collect scores for this specific group
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                _, stored_chat_id, user_id, username, score, _ = row
                if str(stored_chat_id) == chat_id:
                    total_users.add(user_id)  # Count unique users in this group
                    if user_id in group_scores:
                        group_scores[user_id]["score"] += int(score)
                    else:
                        group_scores[user_id] = {"username": escape_markdown(username), "score": int(score)}

        if not group_scores:
            await update.message.reply_text("No scores available for this group.")
            return

        # Sort and get top 10
        sorted_group_scores = sorted(group_scores.items(), key=lambda x: x[1]["score"], reverse=True)
        leaderboard = f"🏆 *Top 10 Scorers in This Group* 🏆\n👥 *Total Members in Leaderboard:* `{len(total_users)}`\n\n"
        
        for rank, (user_id, data) in enumerate(sorted_group_scores[:10], start=1):
            leaderboard += f"{rank}\\) *{data['username']}* \\- `{data['score']} points`\n"

        await update.message.reply_text(leaderboard, parse_mode="MarkdownV2")

    except Exception as e:
        print(f"Error in top1grp_scorer: {e}")

def escape_markdown(text):
    """Escape special characters for Telegram MarkdownV2."""
    escape_chars = r"_*[]()~`>#+-=|{}.!\\"
    return re.sub(r"([{}])".format(re.escape(escape_chars)), r"\\\1", text)

async def all_time_topper(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Returns the top 10 scorers across all groups, including total participants."""
    try:
        # Load Excel file
        try:
            workbook = load_workbook(SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            await update.message.reply_text("No scores found.")
            return

        global_scores = {}
        total_users = set()

        # Collect scores for all users across all groups
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                _, chat_id, user_id, username, score, _ = row
                total_users.add(user_id)  # Count unique users across all groups
                if user_id in global_scores:
                    global_scores[user_id]["score"] += int(score)
                else:
                    global_scores[user_id] = {"username": escape_markdown(username), "score": int(score)}

        if not global_scores:
            await update.message.reply_text("No scores available.")
            return

        # Sort and get top 10
        sorted_global_scores = sorted(global_scores.items(), key=lambda x: x[1]["score"], reverse=True)
        leaderboard = f"🌍 *All\\-Time Top 10 Scorers* 🌍\n👥 *Total Participants:* `{len(total_users)}`\n\n"
        
        for rank, (user_id, data) in enumerate(sorted_global_scores[:10], start=1):
            leaderboard += f"{rank}\\) *{data['username']}* \\- `{data['score']} points`\n"

        await update.message.reply_text(leaderboard, parse_mode="MarkdownV2")

    except Exception as e:
        print(f"Error in all_time_topper: {e}")


def main():
    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler('startquiz', start_game_command))
    application.add_handler(CallbackQueryHandler(handle_difficulty_selection, pattern='^difficulty_'))
    application.add_handler(CallbackQueryHandler(handle_type_selection, pattern='^type_'))
    application.add_handler(CallbackQueryHandler(handle_time_selection, pattern='^time_'))
    application.add_handler(CallbackQueryHandler(handle_button_click, pattern=r'^\d+$'))
    application.add_handler(PollAnswerHandler(handle_poll_answer))
    application.add_handler(CommandHandler("myrank", my_rank))
    application.add_handler(CommandHandler("topgrpscorer", top1grp_scorer))
    application.add_handler(CommandHandler("alltimetopper", all_time_topper))
    application.add_handler(CommandHandler('cancelquiz', cancel_quiz_command))

    # Start the bot
    application.run_polling()

if __name__ == '__main__':
    main()
